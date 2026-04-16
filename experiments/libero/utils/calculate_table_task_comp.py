"""
calculate_table_task_comp.py

Genera una tabella Excel con i risultati di Task Composition L1 o L2.

Colonne:
  - # Task
  - Task Command (comando testato)
  - Reference Task (task di riferimento dal training)
  - SR% per seed 0, 1, 2
  - Task Completion per seed 0, 1, 2
  - Mean SR% ± Std
  - Mean Task Completion

Uso:
  python calculate_table_task_comp.py --txt_dir <dir> output.xlsx
  python calculate_table_task_comp.py --level l2 --txt_dir <dir> output.xlsx
  python calculate_table_task_comp.py --manual --seed0 f0.txt --seed1 f1.txt --seed2 f2.txt output.xlsx
"""

import re
import math
import argparse
import os
import glob
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# =========================
# TASK ORDER & MAPPING
# =========================

def get_task_comp_l1_order():
    """Ordine fisso dei 5 task composition L1."""
    return [
        "Put the plate on the top of the cabinet",
        "Put the plate on the stove",
        "Put the cream cheese on the top of the cabinet",
        "Put the cream cheese on the plate",
        "Open the top layer of the drawer and put the cream cheese inside",
    ]


def get_task_comp_l2_order():
    """Ordine fisso dei 5 task composition L2."""
    return [
        "Open the middle layer of the drawer and put the bowl inside",
        "Put the bowl on the stove and turn on the stove",
        "Put the cream cheese on the bowl and put the bowl on the plate",
        "Push the plate to the front of the stove and put the bowl on the plate",
        "Put the cream cheese on the bowl and put the bowl on the top of the cabinet",
    ]


def get_reference_mapping(level="l1"):
    """
    Mapping: task composition → task di riferimento dal training.
    Chiavi lowercase per matching case-insensitive.
    """
    if level == "l2":
        return {
            "open the middle layer of the drawer and put the bowl inside":
                "Open the middle drawer of the cabinet  /  Open the top drawer and put the bowl inside",
            "put the bowl on the stove and turn on the stove":
                "Put the bowl on the stove  /  Turn on the stove",
            "put the cream cheese on the bowl and put the bowl on the plate":
                "Put the cream cheese in the bowl  /  Put the bowl on the plate",
            "push the plate to the front of the stove and put the bowl on the plate":
                "Push the plate to the front of the stove  /  Put the bowl on the plate",
            "put the cream cheese on the bowl and put the bowl on the top of the cabinet":
                "Put the cream cheese in the bowl  /  Put the bowl on the top of the cabinet",
        }
    return {
        "put the plate on the top of the cabinet":
            "Put the bowl on the top of the cabinet",
        "put the plate on the stove":
            "Put the bowl on the stove",
        "put the cream cheese on the top of the cabinet":
            "Put the wine bottle on the top of the cabinet",
        "put the cream cheese on the plate":
            "Put the cream cheese on the bowl",
        "open the top layer of the drawer and put the cream cheese inside":
            "Open the top layer of the drawer and put the bowl inside",
    }


# =========================
# PARSING
# =========================

def parse_txt_file(filepath):
    """
    Parsa un file di log task_comp_l1 (OpenVLA / TinyVLA).
    Cerca blocchi TASK X/N con "Command:" e "Task success rate:".

    Returns:
        rates:  {task_command_lower: success_rate_%}
        episodes: {task_command_lower: num_episodes}
        task_order: [task_command (originale case)]
    """
    rates = {}
    episodes = {}
    task_order = []

    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    current_task = None
    current_task_lower = None
    current_episodes = None

    for line in lines:
        line_s = line.strip()

        # "Command: Put the plate on the stove"
        if line_s.startswith("Command:"):
            current_task = line_s.split("Command:", 1)[-1].strip()
            current_task_lower = current_task.lower()

        # "# episodes completed: 50"  (ultimo valore = totale)
        if current_task and "# episodes completed:" in line_s:
            m = re.search(r"# episodes completed:\s*(\d+)", line_s)
            if m:
                current_episodes = int(m.group(1))

        # "# episodes: 50"  (alternativa TinyVLA)
        if current_task and "# episodes:" in line_s and "completed" not in line_s:
            m = re.search(r"# episodes:\s*(\d+)", line_s)
            if m:
                current_episodes = int(m.group(1))

        # "Task success rate: 0.0000 ..."
        if current_task and "Task success rate:" in line_s:
            m = re.search(r"Task success rate:\s*([\d.]+)", line_s)
            if m:
                rate_pct = float(m.group(1)) * 100
                if current_task_lower not in rates:
                    rates[current_task_lower] = rate_pct
                    episodes[current_task_lower] = current_episodes if current_episodes else 50
                    task_order.append(current_task)
                current_task = None
                current_task_lower = None
                current_episodes = None

    return rates, episodes, task_order


def merge_txt_files(txt_files_list):
    """Combina più file .txt dello stesso seed."""
    task_rates_list = defaultdict(list)
    task_episodes = {}
    all_tasks = []

    for fp in txt_files_list:
        rates, eps, order = parse_txt_file(fp)
        for task in order:
            tk = task.lower()
            if tk not in task_rates_list:
                all_tasks.append(task)
            task_rates_list[tk].append(rates[tk])
            task_episodes[tk] = eps.get(tk, 50)

    merged_rates = {tk: sum(v) / len(v) for tk, v in task_rates_list.items()}
    return merged_rates, task_episodes, all_tasks


# =========================
# EXCEL GENERATION
# =========================

def write_excel(output_xlsx, txt_files_by_seed, level="l1"):
    """Genera l'Excel con i risultati task_comp (L1 o L2)."""

    # Parse per ogni seed
    all_rates = []
    all_episodes = []
    for seed_idx in range(3):
        fps = txt_files_by_seed[seed_idx]
        print(f"  Seed {seed_idx}: {len(fps)} file")
        for fp in fps:
            print(f"    - {os.path.basename(fp)}")
        rates, eps, _ = merge_txt_files(fps)
        all_rates.append(rates)
        all_episodes.append(eps)
        print(f"    -> {len(rates)} task trovati")

    ref_mapping = get_reference_mapping(level)
    fixed_order = get_task_comp_l2_order() if level == "l2" else get_task_comp_l1_order()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Task Comp {level.upper()}"

    # Header
    headers = [
        "# Task",
        "Task Command",
        "Reference Task",
        "SR% - Seed 0", "Task Completion - Seed 0",
        "SR% - Seed 1", "Task Completion - Seed 1",
        "SR% - Seed 2", "Task Completion - Seed 2",
        "Mean SR% ± Std",
        "Mean Task Completion",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Accumula per riga finale
    all_seed_rates = [[], [], []]
    all_seed_completions = [[], [], []]

    for task_idx, task_cmd in enumerate(fixed_order, start=1):
        tk = task_cmd.lower()
        ref_task = ref_mapping.get(tk, "")

        seed_rates = []
        seed_completions = []

        for seed_idx in range(3):
            rate = all_rates[seed_idx].get(tk, float("nan"))
            succ = int(round(rate / 100 * 50)) if rate == rate else 0

            seed_rates.append(rate)
            seed_completions.append(f"{succ}/50")

            if rate == rate:
                all_seed_rates[seed_idx].append(rate)
                all_seed_completions[seed_idx].append((succ, 50))

        # Mean ± Std
        valid = [r for r in seed_rates if r == r]
        if valid:
            mean_r = sum(valid) / len(valid)
            std_r = math.sqrt(sum((r - mean_r) ** 2 for r in valid) / (len(valid) - 1)) if len(valid) > 1 else 0.0
            mean_display = f"{mean_r:.1f}% ± {std_r:.1f}%"
            avg_succ = int(round(mean_r / 100 * 50))
        else:
            mean_display = "N/A"
            avg_succ = 0

        ws.append([
            task_idx,
            task_cmd,
            ref_task,
            f"{seed_rates[0]:.1f}%" if seed_rates[0] == seed_rates[0] else "N/A",
            seed_completions[0],
            f"{seed_rates[1]:.1f}%" if seed_rates[1] == seed_rates[1] else "N/A",
            seed_completions[1],
            f"{seed_rates[2]:.1f}%" if seed_rates[2] == seed_rates[2] else "N/A",
            seed_completions[2],
            mean_display,
            f"{avg_succ}/50",
        ])

    # Riga finale: Mean ± Std globale
    final_row = ["", "Mean% ± Std%", ""]
    for seed_idx in range(3):
        rates = all_seed_rates[seed_idx]
        if rates:
            m = sum(rates) / len(rates)
            s = math.sqrt(sum((r - m) ** 2 for r in rates) / (len(rates) - 1)) if len(rates) > 1 else 0.0
            ts = sum(c[0] for c in all_seed_completions[seed_idx])
            te = sum(c[1] for c in all_seed_completions[seed_idx])
            final_row.append(f"{m:.2f}% ± {s:.2f}%")
            final_row.append(f"{ts}/{te}")
        else:
            final_row.extend(["N/A", "0/0"])

    # Media globale tra seed
    seed_means = []
    for seed_idx in range(3):
        if all_seed_rates[seed_idx]:
            seed_means.append(sum(all_seed_rates[seed_idx]) / len(all_seed_rates[seed_idx]))
    if seed_means:
        gm = sum(seed_means) / len(seed_means)
        gs = math.sqrt(sum((m - gm) ** 2 for m in seed_means) / (len(seed_means) - 1)) if len(seed_means) > 1 else 0.0
        gt_s = sum(sum(c[0] for c in all_seed_completions[i]) for i in range(3))
        gt_e = sum(sum(c[1] for c in all_seed_completions[i]) for i in range(3))
        final_row.append(f"{gm:.2f}% ± {gs:.2f}%")
        final_row.append(f"{gt_s}/{gt_e}")
    else:
        final_row.extend(["N/A", "0/0"])

    ws.append(final_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len + 2, 65)

    wb.save(output_xlsx)
    print(f"\n[OK] Excel salvato: {output_xlsx}")


# =========================
# AUTO-DETECT
# =========================

def find_txt_files_by_seed(base_dir, pattern_prefix=None, level="l1"):
    """Trova i file .txt per ogni seed nella directory."""
    if pattern_prefix is None:
        pattern_prefix = f"EVAL-task_comp_{level}"
    txt_files_by_seed = {0: [], 1: [], 2: []}

    print(f"\n[INFO] Cercando in: {base_dir}")
    for seed_idx in range(3):
        pattern = os.path.join(base_dir, f"{pattern_prefix}*seed_{seed_idx}*.txt")
        matches = sorted(glob.glob(pattern))
        # Fallback: pattern senza underscore prima del numero
        if not matches:
            pattern = os.path.join(base_dir, f"{pattern_prefix}*seed{seed_idx}*.txt")
            matches = sorted(glob.glob(pattern))
        txt_files_by_seed[seed_idx] = matches
        if matches:
            print(f"  Seed {seed_idx}: {len(matches)} file")
            for m in matches:
                print(f"    - {os.path.basename(m)}")
        else:
            print(f"  Seed {seed_idx}: NESSUN FILE!")
            return None

    return txt_files_by_seed


# =========================
# MAIN
# =========================

def main():
    parser = argparse.ArgumentParser(
        description="Genera tabella Excel Task Composition L1/L2 da file log .txt"
    )

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--txt_dir", help="Directory con i file .txt (auto-detect)")
    group.add_argument("--manual", action="store_true", help="Specifica file manualmente")

    parser.add_argument("--level", default="l1", choices=["l1", "l2"],
                        help="Livello di composizione: l1 (default) o l2")
    parser.add_argument("--pattern", default=None,
                        help="Prefix per auto-detect (default: EVAL-task_comp_<level>)")
    parser.add_argument("--seed0", nargs="+", help="File per seed 0 (modo manuale)")
    parser.add_argument("--seed1", nargs="+", help="File per seed 1 (modo manuale)")
    parser.add_argument("--seed2", nargs="+", help="File per seed 2 (modo manuale)")
    parser.add_argument("output_xlsx", help="File Excel di output")

    args = parser.parse_args()

    print("\n" + "=" * 60)
    print(f"  TASK COMPOSITION {args.level.upper()} - EXCEL TABLE GENERATOR")
    print("=" * 60)

    if args.manual:
        if not (args.seed0 and args.seed1 and args.seed2):
            print("[ERROR] --manual richiede --seed0, --seed1, --seed2")
            exit(1)
        txt_files_by_seed = {0: args.seed0, 1: args.seed1, 2: args.seed2}
    else:
        txt_files_by_seed = find_txt_files_by_seed(args.txt_dir, args.pattern, args.level)
        if txt_files_by_seed is None:
            print("[ERROR] File non trovati per tutti i seed!")
            exit(1)

    write_excel(args.output_xlsx, txt_files_by_seed, level=args.level)

    print("=" * 60)
    print("  COMPLETATO!")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
