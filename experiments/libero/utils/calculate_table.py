import re
import math
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# =========================
# MAPPING COMPLETO DELLE VARIAZIONI
# =========================

def get_variation_mapping():
    """
    Mapping completo: ogni variazione → task originale
    Include: DEFAULT, L1, L2, L3
    """
    mapping = {}
    
    # Task 1: Open the middle layer of the drawer
    orig_1 = "Open the middle layer of the drawer"
    mapping["Open the middle layer of the drawer"] = orig_1  # DEFAULT
    mapping["Open the middle layer of the cabinet"] = orig_1  # DEFAULT VARIANT
    mapping["Pull the middle layer of the drawer"] = orig_1  # L1
    mapping["The middle layer of the drawer needs to be opened"] = orig_1  # L2
    mapping["Open the layer of the drawer located between the top and bottom"] = orig_1  # L3
    
    # Task 2: Put the bowl on the stove
    orig_2 = "Put the bowl on the stove"
    mapping["Put the bowl on the stove"] = orig_2  # DEFAULT
    mapping["Set the bowl on the stove"] = orig_2  # L1
    mapping["The stove needs to have the bowl on it"] = orig_2  # L2
    mapping["Put the object between the wine bottle and the cream cheese on the stove"] = orig_2  # L3
    
    # Task 3: Put the wine bottle on the top of the drawer
    orig_3 = "Put the wine bottle on the top of the drawer"
    mapping["Put the wine bottle on top of the drawer"] = orig_3  # DEFAULT
    mapping["Put the wine bottle on the top of the cabinet"] = orig_3  # DEFAULT VARIANT
    mapping["Place the wine bottle on the top of the cabinet"] = orig_3  # L1
    mapping["Top of the cabinet needs to have the wine bottle on it"] = orig_3  # L2
    mapping["Put the object behind the bowl on the top of the cabinet"] = orig_3  # L3
    
    # Task 4: Open the top layer of the drawer and put the bowl inside
    orig_4 = "Open the top layer of the drawer and put the bowl inside"
    mapping["Open the top drawer and put the bowl inside"] = orig_4  # DEFAULT
    mapping["Open the top layer of the drawer and put the bowl inside"] = orig_4  # DEFAULT VARIANT
    mapping["Pull the top layer of the drawer and place the bowl inside"] = orig_4  # L1
    mapping["Pull the top layer of the drawer and put the bowl inside"] = orig_4  # L1 (variant)
    mapping["The top layer of the drawer needs to be opened and the bowl needs to be put inside"] = orig_4  # L2
    mapping["Open the top layer of the drawer and put the object between the plate and the cream cheese inside"] = orig_4  # L3
    
    # Task 5: Put the bowl on the top of the drawer
    orig_5 = "Put the bowl on the top of the drawer"
    mapping["Put the bowl on top of the drawer"] = orig_5  # DEFAULT
    mapping["Put the bowl on the top of the cabinet"] = orig_5  # DEFAULT VARIANT
    mapping["Place the bowl on the top of the cabinet"] = orig_5  # L1
    mapping["The top of the cabinet needs to have the bowl on it"] = orig_5  # L2
    mapping["Put the object between the wine bottle and the cream cheese on the top of the cabinet"] = orig_5  # L3
    
    # Task 6: Push the plate to the front of the stove
    orig_6 = "Push the plate to the front of the stove"
    mapping["Push the plate to the front of the stove"] = orig_6  # DEFAULT
    mapping["Move the plate to the front of the stove"] = orig_6  # L1
    mapping["The space in front of the stove needs to have the plate in it"] = orig_6  # L2
    mapping["Push the object in front of the drawer to the front of the stove"] = orig_6  # L3
    
    # Task 7: Put the cream cheese in the bowl
    orig_7 = "Put the cream cheese in the bowl"
    mapping["Put the cream cheese in the bowl"] = orig_7  # DEFAULT
    mapping["Put the cream cheese on the bowl"] = orig_7  # DEFAULT VARIANT
    mapping["Place the cream cheese on the bowl"] = orig_7  # L1
    mapping["Place the cream cheese in the bowl"] = orig_7  # L1 (variant)
    mapping["The cream cheese needs to be put on the bowl"] = orig_7  # L2
    mapping["Put the object in front of the stove on the bowl"] = orig_7  # L3
    
    # Task 8: Turn on the stove
    orig_8 = "Turn on the stove"
    mapping["Turn on the stove"] = orig_8  # DEFAULT
    mapping["Switch on the stove"] = orig_8  # L1
    mapping["The stove needs to be turned on"] = orig_8  # L2
    mapping["Turn on the object behind the cream cheese"] = orig_8  # L3
    
    # Task 9: Put the bowl on the plate
    orig_9 = "Put the bowl on the plate"
    mapping["Put the bowl on the plate"] = orig_9  # DEFAULT
    mapping["Place the bowl on the plate"] = orig_9  # L1
    mapping["The plate needs to have the bowl on it"] = orig_9  # L2
    mapping["Put the object between the wine bottle and the cream cheese on the plate"] = orig_9  # L3
    
    # Task 10: Put the wine bottle on the rack
    orig_10 = "Put the wine bottle on the rack"
    mapping["Put the wine bottle on the rack"] = orig_10  # DEFAULT
    mapping["Place the wine bottle on the rack"] = orig_10  # L1
    mapping["The rack needs to be filled with the wine bottle in it"] = orig_10  # L2
    mapping["Put the object behind the bowl on the rack"] = orig_10  # L3
    
    return mapping


# =========================
# PARSING TXT FILES
# =========================

def parse_txt_file(filepath):
    """
    Parsa il file txt e estrae task, success rate ed episodi
    Mantiene l'ordine in cui appaiono i task
    """
    rates = {}
    episodes = {}
    task_order = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    in_table = False
    for line in lines:
        line = line.strip()
        
        # Inizia la tabella dopo "Task | Success Rate | Episodes"
        if "Task" in line and "Success Rate" in line and "Episodes" in line:
            in_table = True
            continue
        
        # Fine tabella
        if line.startswith("----") and in_table:
            continue
        
        if "OVERALL" in line and in_table:
            # Parsa la riga OVERALL
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                rate_str = parts[1].replace("%", "").strip()
                ep_str = parts[2].strip()
                rates["OVERALL"] = float(rate_str)
                episodes["OVERALL"] = int(ep_str)
            break
        
        if in_table and "|" in line and not line.startswith("="):
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                task = parts[0].strip()
                rate_str = parts[1].replace("%", "").strip()
                ep_str = parts[2].strip()
                
                if task and rate_str and ep_str:
                    try:
                        rates[task] = float(rate_str)
                        episodes[task] = int(ep_str)
                        task_order.append(task)
                    except ValueError:
                        pass
    
    return rates, episodes, task_order


# =========================
# EXCEL GENERATION
# =========================

def write_excel_comparison(output_xlsx, txt_files):
    """
    Genera l'Excel con il formato richiesto
    """
    # Parsa i 3 file
    all_rates = []
    all_episodes = []
    task_order = None
    
    print("\n[INFO] Parsing dei file txt...")
    for i, filepath in enumerate(txt_files):
        print(f"  - Seed {i}: {filepath}")
        rates, episodes, order = parse_txt_file(filepath)
        all_rates.append(rates)
        all_episodes.append(episodes)
        if task_order is None:
            task_order = order  # Usa l'ordine del primo file
        print(f"    ✓ Trovati {len(rates)-1} task (+ OVERALL)")
    
    # Ottieni il mapping variazione → originale
    variation_to_original = get_variation_mapping()
    
    # Crea lista di task originali nell'ordine del file txt
    orig_tasks_ordered = []
    variation_tasks_ordered = []
    
    print("\n[INFO] Identificazione variazioni nei file...")
    for task in task_order:
        if task in variation_to_original:
            orig = variation_to_original[task]
            orig_tasks_ordered.append(orig)
            variation_tasks_ordered.append(task)
            print(f"  ✓ '{task}' → '{orig}'")
        else:
            print(f"  ⚠ Non mappata: '{task}'")
            orig_tasks_ordered.append(f"UNKNOWN: {task}")
            variation_tasks_ordered.append(task)
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    # Header
    ws.append([
        "Original Task Command",
        "Variation Task Command",
        "Success Rate (%) - Seed 0",
        "Task Completion - Seed 0",
        "Success Rate (%) - Seed 1",
        "Task Completion - Seed 1",
        "Success Rate (%) - Seed 2",
        "Task Completion - Seed 2",
        "Mean Success Rate (%) ± Std",
        "Mean Task Completion"
    ])
    
    # Formatta header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Statistiche per la riga finale
    all_seed_rates = [[], [], []]
    all_seed_completions = [[], [], []]
    
    print("\n[INFO] Generazione tabella Excel...")
    
    # Processa ogni task NELL'ORDINE DEL FILE TXT
    for orig_task, variation in zip(orig_tasks_ordered, variation_tasks_ordered):
        
        seed_rates = []
        seed_completions = []
        seed_success_counts = []
        seed_episode_counts = []
        
        # Estrai dati per ogni seed
        for seed_idx in range(3):
            if variation in all_rates[seed_idx]:
                rate = all_rates[seed_idx][variation]
                ep = all_episodes[seed_idx][variation]
                
                # Usa 50 episodi reali
                success_count = int(round(rate / 100 * ep))
                
                seed_rates.append(rate)
                seed_completions.append(f"{success_count}/{ep}")
                seed_success_counts.append(success_count)
                seed_episode_counts.append(ep)
                
                all_seed_rates[seed_idx].append(rate)
                all_seed_completions[seed_idx].append((success_count, ep))
            else:
                seed_rates.append(float('nan'))
                seed_completions.append("0/50")
                seed_success_counts.append(0)
                seed_episode_counts.append(50)
        
        # Calcola media E deviazione standard per il task
        valid_rates = [r for r in seed_rates if r == r]  # Escludi NaN
        if valid_rates:
            mean_rate = sum(valid_rates) / len(valid_rates)
            # Calcola std dev tra i seed
            if len(valid_rates) > 1:
                std_rate = math.sqrt(sum((r - mean_rate) ** 2 for r in valid_rates) / (len(valid_rates) - 1))
            else:
                std_rate = 0.0
            mean_std_display = f"{mean_rate:.1f}% ± {std_rate:.1f}%"
        else:
            mean_rate = float('nan')
            std_rate = float('nan')
            mean_std_display = "nan%"
        
        total_success = sum(seed_success_counts)
        total_episodes = sum(seed_episode_counts)
        
        var_display = variation if len(variation) <= 50 else variation[:47] + "..."
        
        # Aggiungi riga
        ws.append([
            orig_task,
            var_display,
            f"{seed_rates[0]:.1f}%" if seed_rates[0] == seed_rates[0] else "nan%",
            seed_completions[0],
            f"{seed_rates[1]:.1f}%" if seed_rates[1] == seed_rates[1] else "nan%",
            seed_completions[1],
            f"{seed_rates[2]:.1f}%" if seed_rates[2] == seed_rates[2] else "nan%",
            seed_completions[2],
            mean_std_display,
            f"{total_success}/{total_episodes}"
        ])
    
    # ================= RIGA FINALE: Mean% ± Std% =================
    final_row = ["Mean% ± Std%", ""]
    
    # Per ogni seed
    for seed_idx in range(3):
        rates = all_seed_rates[seed_idx]
        
        if rates:
            mean = sum(rates) / len(rates)
            if len(rates) > 1:
                std = math.sqrt(sum((r - mean) ** 2 for r in rates) / (len(rates) - 1))
            else:
                std = 0.0
            
            # Somma completion
            total_success = sum(c[0] for c in all_seed_completions[seed_idx])
            total_episodes = sum(c[1] for c in all_seed_completions[seed_idx])
            
            final_row.append(f"{mean:.2f}% ± {std:.2f}%")
            final_row.append(f"{total_success}/{total_episodes}")
        else:
            final_row.append("N/A")
            final_row.append("0/0")
    
    # Media globale (standard VLA: std tra le medie dei seed)
    seed_means = []
    for seed_idx in range(3):
        if all_seed_rates[seed_idx]:
            seed_mean = sum(all_seed_rates[seed_idx]) / len(all_seed_rates[seed_idx])
            seed_means.append(seed_mean)
    
    if seed_means:
        global_mean = sum(seed_means) / len(seed_means)
        if len(seed_means) > 1:
            global_std = math.sqrt(
                sum((m - global_mean) ** 2 for m in seed_means) / (len(seed_means) - 1)
            )
        else:
            global_std = 0.0
        
        global_success = sum(sum(c[0] for c in all_seed_completions[i]) for i in range(3))
        global_episodes = sum(sum(c[1] for c in all_seed_completions[i]) for i in range(3))
        
        final_row.append(f"{global_mean:.2f}% ± {global_std:.2f}%")
        final_row.append(f"{global_success}/{global_episodes}")
    else:
        final_row.append("N/A")
        final_row.append("0/0")
    
    ws.append(final_row)
    
    # Formatta ultima riga
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
    
    # Auto-adjust colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_xlsx)
    print(f"\n[OK] Excel salvato in: {output_xlsx}\n")


# =========================
# MAIN
# =========================

def main():
    parser = argparse.ArgumentParser(description="Genera tabella Excel da file txt")
    parser.add_argument("txt_files", nargs=3, help="3 file txt (seed 0, 1, 2)")
    parser.add_argument("output_xlsx", help="File Excel di output")
    args = parser.parse_args()
    
    print("\n" + "="*60)
    print("  LIBERO TASK COMPARISON FROM TXT FILES")
    print("  (Supporting DEFAULT/L1/L2/L3 Variations)")
    print("="*60)
    
    write_excel_comparison(args.output_xlsx, args.txt_files)
    
    print("="*60)
    print("  COMPLETATO!")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()