"""
analyze_embedding_overlap_rollout.py
=====================================

Quantifies the semantic and lexical distance between the **default** natural-
language instruction of a LIBERO-Goal task and each of its three paraphrase
variants (L1, L2, L3) by comparing VLA hidden-state embeddings extracted
during real policy rollouts.

Scientific Motivation
---------------------
Evaluating how sensitive a Vision-Language-Action (VLA) policy is to
instruction paraphrasing requires more than a simple string comparison. Two
phrases can share few characters yet carry identical meaning (low Levenshtein,
high cosine similarity), or share many words but differ in intent (high
Levenshtein, still high cosine similarity). This script therefore computes
**three complementary metrics** for each default → variant pair:

1. **Cosine Similarity** (semantic, high = similar)
   Measures the angle between two mean embedding vectors in the VLA's internal
   representation space. Values close to 1.0 indicate that the model encodes
   the two instructions in nearly identical directions, suggesting the policy
   "sees" them as semantically equivalent and is therefore likely to produce
   similar action distributions [web:50][web:53].

2. **Euclidean Distance / L2 norm** (semantic, low = similar)
   Measures the absolute displacement between embedding vectors. Unlike cosine
   similarity, Euclidean distance is sensitive to magnitude differences, making
   it a complementary check: two embeddings may be collinear (cos ≈ 1) but
   differ substantially in magnitude, hinting at confidence or activation-scale
   differences.

3. **Normalized Levenshtein Distance** (lexical, low = similar)
   Counts the minimum number of single-character insertions, deletions, and
   substitutions needed to transform one string into the other, normalised by
   the length of the longer string [web:48]. A pure surface-form metric with no
   semantic awareness — its primary use here is to confirm that the paraphrase
   levels (L1 < L2 < L3) represent genuinely increasing lexical divergence
   from the default instruction.

Embedding Extraction Modes
--------------------------
Embeddings are pre-computed by a companion script (``collect_rollout_embeddings.py``)
and serialised as pickle files. Two extraction strategies are supported:

- **full_rollout mode**: the embedding stored for a given instruction is the
  **mean** of all hidden-state vectors collected across every step of every
  rollout. This produces a single representative point per instruction in the
  VLA's representation space, averaging out step-to-step variation and giving a
  stable estimate of the model's "internal concept" for that command.
- **first_step_only mode**: only the hidden-state vector at step 0 of each
  rollout is retained before averaging. This isolates the model's *prior*
  representation of the instruction before any visual feedback is incorporated,
  which can reveal whether the language tower alone distinguishes variants.

Input Data Format
-----------------
Each pickle file is a ``dict[str, dict]`` keyed by strings of the form
``"task_{N:02d}_{level}"`` (e.g. ``"task_07_default"``, ``"task_07_l1"``).
Each value is a metadata dict with at minimum the following keys:

+--------------------------------+---------------------------------------------+
| Key                            | Description                                 |
+================================+=============================================+
| ``"embedding"``                | ``np.ndarray`` — mean embedding vector over |
|                                | all rollout steps (shape: ``(D,)`` or       |
|                                | ``(1, D)`` where D is LLM hidden dim)       |
+--------------------------------+---------------------------------------------+
| ``"embedding_per_rollout"``    | ``np.ndarray`` — one embedding per rollout  |
|                                | (shape: ``(R, D)`` where R = num_rollouts)  |
+--------------------------------+---------------------------------------------+
| ``"command_text"``             | ``str`` — the instruction string used       |
+--------------------------------+---------------------------------------------+
| ``"command_level"``            | ``str`` — one of ``"default"``, ``"l1"``,   |
|                                | ``"l2"``, ``"l3"``                          |
+--------------------------------+---------------------------------------------+
| ``"task_id"``                  | ``int`` — 0-indexed LIBERO-Goal task ID     |
+--------------------------------+---------------------------------------------+
| ``"num_rollouts"``             | ``int`` (optional) — number of rollouts     |
+--------------------------------+---------------------------------------------+
| ``"success_rate"``             | ``float`` (optional) — policy success rate  |
+--------------------------------+---------------------------------------------+
| ``"first_step_only"``          | ``bool`` (optional) — extraction mode flag  |
+--------------------------------+---------------------------------------------+
| ``"total_steps"``              | ``int`` (optional) — total steps collected  |
+--------------------------------+---------------------------------------------+

Output Files
------------
For each run, the script produces:

- ``{output_base}.csv``               — full results table (one row per pair)
- ``{output_base}_overlap_l1.xlsx``   — formatted Excel table for L1 variants
- ``{output_base}_overlap_l2.xlsx``   — formatted Excel table for L2 variants
- ``{output_base}_overlap_l3.xlsx``   — formatted Excel table for L3 variants

Typical CLI Usage
-----------------
    # Single file (backwards-compatible)
    python analyze_embedding_overlap_rollout.py --embedding_file embeddings.pkl

    # Multiple files, one per level
    python analyze_embedding_overlap_rollout.py --embedding_files \\
        .../rollout_embeddings_libero_goal_default_first_step_r10.pkl \\
        .../rollout_embeddings_libero_goal_l1_first_step_r10.pkl \\
        .../rollout_embeddings_libero_goal_l2_first_step_r10.pkl \\
        .../rollout_embeddings_libero_goal_l3_first_step_r10.pkl

    # Per-rollout granular analysis
    python analyze_embedding_overlap_rollout.py \\
        --embedding_files ... --per_rollout_analysis --task_id 7 --level l2

Dependencies
------------
- numpy              : Array operations, statistics
- scikit-learn       : cosine_similarity, euclidean_distances
- pandas             : DataFrame construction and CSV export
- python-Levenshtein : Fast C-extension Levenshtein distance computation
- openpyxl           : Excel file generation with conditional styling

Author: Agostino Cardamone
"""


# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================

import os      # Path operations: os.path.join, os.path.dirname, os.path.basename
import glob    # Unix-style wildcard file discovery: glob.glob("*.pkl")
import pickle  # Deserialise Python objects from binary .pkl files


# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================

import numpy as np   # Numerical arrays: embedding stacking, mean, std, min, max

from sklearn.metrics.pairwise import (
    cosine_similarity,    # Computes cos(θ) ∈ [-1, 1] between pairs of row vectors.
                          # Returns an (m, n) matrix when given (m, D) and (n, D) inputs.
    euclidean_distances,  # Computes L2 ‖a − b‖ between pairs of row vectors.
                          # Returns an (m, n) matrix when given (m, D) and (n, D) inputs.
)

import pandas as pd  # Tabular data: DataFrame construction, groupby, CSV/Excel export

from Levenshtein import distance as levenshtein_distance  # Fast Cython-level string edit
                                                           # distance (python-Levenshtein).
                                                           # Returns the raw integer count
                                                           # of minimum single-char edits.


# =============================================================================
# NORMALISED LEVENSHTEIN DISTANCE
# =============================================================================

def compute_levenshtein_normalized(text1: str, text2: str) -> float:
    """
    Compute the normalised Levenshtein (edit) distance between two strings.

    The raw Levenshtein distance counts the minimum number of single-character
    insertions, deletions, or substitutions required to transform ``text1``
    into ``text2`` [web:48]. Because longer strings naturally incur higher raw
    distances, the value is divided by the length of the longer string to
    produce a normalised score in ``[0, 1]`` that is comparable across pairs
    of instructions with different lengths.

    Normalisation Formula
    ---------------------
    Let ``d = levenshtein(text1, text2)`` and ``m = max(len(text1), len(text2))``.
    Then:

    .. math::

        \\text{NLD}(text1, text2) = \\frac{d}{m}

    Interpretation
    --------------
    - ``0.0`` — strings are identical (zero edits needed).
    - ``1.0`` — strings share no characters; one is a complete rewrite of the
      other (every character must be substituted or deleted/inserted).
    - Typical paraphrase levels in this study:
      - L1 (minor synonym substitutions): NLD ≈ 0.05–0.20
      - L2 (structural rephrasing):       NLD ≈ 0.20–0.50
      - L3 (complete reformulation):      NLD ≈ 0.50–0.90

    Edge Cases
    ----------
    If both strings are empty (``max_len == 0``), the function returns ``0.0``
    (two empty strings are trivially identical) rather than raising a
    ``ZeroDivisionError``.

    Parameters
    ----------
    text1 : str
        First instruction string (e.g. ``"Turn on the stove"``).
    text2 : str
        Second instruction string (e.g. ``"Activate the stove"``).

    Returns
    -------
    float
        Normalised Levenshtein distance in the range ``[0.0, 1.0]``.
        ``0.0`` = identical strings; ``1.0`` = maximally different.

    Examples
    --------
    >>> compute_levenshtein_normalized("Turn on the stove", "Turn on the stove")
    0.0
    >>> compute_levenshtein_normalized("Turn on the stove", "stove")
    0.7222...
    >>> compute_levenshtein_normalized("", "")
    0.0
    """
    # Compute the raw integer edit distance between the two strings.
    # Uses the fast python-Levenshtein C extension for O(m*n) time complexity.
    lev_dist = levenshtein_distance(text1, text2)

    # Determine the normalisation denominator: length of the longer string.
    # Normalising by the longer string ensures the score stays in [0, 1] even
    # when one string is much shorter (e.g. "stove" vs "Turn on the stove").
    max_len = max(len(text1), len(text2))

    # Guard: if both strings are empty, return 0 (identical) without dividing.
    if max_len == 0:
        return 0.0

    # Divide raw distance by max length to normalise to [0, 1].
    return lev_dist / max_len


# =============================================================================
# MULTI-FILE EMBEDDING LOADER
# =============================================================================

def load_embeddings(
    embedding_files: list = None,
    embedding_dir: str = None,
) -> dict:
    """
    Load and merge embedding dictionaries from multiple pickle files.

    Supports two source modes:

    1. **Directory mode** (``embedding_dir`` provided): discovers all ``*.pkl``
       files in the directory using ``glob``, sorted alphabetically so that
       files from the same run are merged in a consistent order.
    2. **File list mode** (``embedding_files`` provided): loads exactly the
       files listed, in the order given.

    All loaded dictionaries are merged into a single flat dict using
    ``dict.update()``. If two files define the same key (e.g. the same
    ``"task_07_default"`` entry), the value from the **later file** in the
    loading order overwrites the earlier one. This is intentional: it allows
    a newer re-computed file to supersede an older one without raising an error.

    Parameters
    ----------
    embedding_files : list of str or None, optional
        Ordered list of absolute or relative paths to ``.pkl`` files.
        Mutually exclusive with ``embedding_dir``.
    embedding_dir : str or None, optional
        Path to a directory. All ``*.pkl`` files found there are loaded.
        Mutually exclusive with ``embedding_files``.

    Returns
    -------
    all_embeddings : dict[str, dict]
        Merged dictionary mapping each ``"task_{N:02d}_{level}"`` key to its
        embedding metadata dict. The total number of entries is the sum of
        unique keys across all loaded files (with later files winning on
        conflicts).

    Raises
    ------
    ValueError
        If neither ``embedding_files`` nor ``embedding_dir`` is provided.

    Examples
    --------
    >>> embs = load_embeddings(embedding_dir="/outputs/embeddings/")
    Found 4 pickle files in /outputs/embeddings/
      Loading: rollout_embeddings_libero_goal_default_first_step_r10.pkl
      Loading: rollout_embeddings_libero_goal_l1_first_step_r10.pkl
      ...
    Total embeddings loaded: 40
    """
    all_embeddings = {}  # Accumulator: all entries from all files are merged here

    # ── Determine which files to load ─────────────────────────────────────────
    if embedding_dir:
        # Discover all .pkl files in the directory; sort for deterministic order
        files = sorted(glob.glob(os.path.join(embedding_dir, "*.pkl")))
        print(f"Found {len(files)} pickle files in {embedding_dir}")
    elif embedding_files:
        # Use the caller-specified list verbatim
        files = embedding_files
    else:
        # Neither source provided: fail with a descriptive error
        raise ValueError("Must provide embedding_files or embedding_dir")

    # ── Load and merge each file into the accumulator ─────────────────────────
    for filepath in files:
        # Print just the filename (not the full path) for concise console output
        print(f"  Loading: {os.path.basename(filepath)}")

        with open(filepath, 'rb') as f:
            # Deserialise the pickle file: expects a dict[str, dict]
            data = pickle.load(f)
            # Merge into the accumulator; later files overwrite duplicate keys
            all_embeddings.update(data)

    # Summary line: helps the user verify that all expected entries were found
    print(f"\nTotal embeddings loaded: {len(all_embeddings)}")

    return all_embeddings


# =============================================================================
# MAIN ANALYSIS FUNCTION
# =============================================================================

def analyze_embeddings(
    embeddings: dict,
    output_csv: str = "analysis_results.csv",
) -> pd.DataFrame:
    """
    Compute and report semantic + lexical distance metrics for all
    default → variant instruction pairs in the loaded embeddings.

    For each LIBERO-Goal task that has a ``"default"`` entry, this function
    compares the default instruction against every variant level (L1, L2, L3)
    using three metrics:

    - **Cosine Similarity** (sklearn): dot product normalised by vector
      magnitudes, giving cos(θ) ∈ [-1, 1]. Higher = more semantically similar.
    - **Euclidean Distance L2** (sklearn): ‖default_emb − variant_emb‖₂.
      Lower = embeddings are geometrically closer in the representation space.
    - **Normalised Levenshtein Distance**: surface-form character-level edit
      distance normalised to [0, 1]. Lower = more lexically similar [web:48].

    Embedding Handling
    ------------------
    Embeddings may be stored as either 1D ``(D,)`` or 2D ``(1, D)`` arrays
    depending on the extraction script. The function calls ``.flatten()`` on
    any embedding with ``ndim > 1`` before computing pairwise distances,
    ensuring both sklearn functions receive plain 1D vectors wrapped in a
    single-item list (producing a ``(1, 1)`` matrix from which ``[0, 0]`` is
    extracted as a scalar).

    Output
    ------
    After computing all pairwise metrics, the function:

    1. Prints a per-task, per-level breakdown to stdout.
    2. Prints per-level and overall aggregate statistics (mean ± std).
    3. Saves the full results as a CSV at ``output_csv``.
    4. Generates three formatted Excel files (one per level: L1, L2, L3) by
       calling ``_save_formatted_overlap_table``.

    Parameters
    ----------
    embeddings : dict[str, dict]
        Pre-loaded embedding dictionary (output of ``load_embeddings`` or a
        raw ``pickle.load`` call). Keys must follow the ``"task_{N:02d}_{level}"``
        naming convention. Each value must contain at minimum:
        ``"embedding"``, ``"task_id"``, ``"command_level"``, ``"command_text"``.
    output_csv : str, default ``"analysis_results.csv"``
        Absolute or relative path for the output CSV file. The three Excel
        files are written alongside it with ``_overlap_{level}.xlsx`` suffixes
        derived by splitting off the ``.csv`` extension.

    Returns
    -------
    df : pd.DataFrame
        One row per (task_id, level) pair with columns:
        ``task_id``, ``level``, ``default_command``, ``variation_command``,
        ``default_rollouts``, ``variation_rollouts``, ``default_success_rate``,
        ``variation_success_rate``, ``cosine_similarity``,
        ``euclidean_distance``, ``levenshtein_distance``.
    """
    print(f"\nAnalyzing {len(embeddings)} rollout-based mean embeddings\n")

    # ── Diagnostic info from the first entry ──────────────────────────────────
    # Prints one representative entry's metadata so the user can verify that
    # the correct pickle files were loaded (right mode, shape, rollout count).
    if embeddings:
        first_key  = next(iter(embeddings.keys()))  # Arbitrary first key
        first_data = embeddings[first_key]

        print(f"  Mean embedding shape:            {first_data['embedding'].shape}")
        print(f"  Embeddings per rollout shape:    {first_data['embedding_per_rollout'].shape}")
        print(f"  Number of rollouts:              {first_data.get('num_rollouts', 'N/A')}")

        # Determine extraction mode from the 'first_step_only' flag
        mode = (
            "First step only"
            if first_data.get('first_step_only', False)
            else "Full rollout"
        )
        print(f"  Mode:                            {mode}")

        # Optional metadata fields; printed only when present
        if 'total_steps' in first_data:
            print(f"  Total steps:                     {first_data['total_steps']}")
        if 'success_rate' in first_data:
            print(f"  Success rate:                    {first_data['success_rate']:.2%}")
        print()

    # ── Organise embeddings by task_id and command_level ──────────────────────
    # tasks[task_id][command_level] = data_dict
    # This nested structure lets us easily retrieve the 'default' entry and
    # each variant level for a given task with a simple dict lookup.
    tasks = {}
    for key, data in embeddings.items():
        task_id = data['task_id']
        if task_id not in tasks:
            tasks[task_id] = {}  # First entry for this task_id
        tasks[task_id][data['command_level']] = data  # Store by level

    results = []  # Will be converted to a DataFrame at the end

    # ── Per-task, per-level metric computation ────────────────────────────────
    print("=" * 80)
    print("DISTANCE ANALYSIS: Semantic (Cosine + Euclidean) + Lexical (Levenshtein)")
    print("Using ROLLOUT MEAN EMBEDDINGS for semantic similarity")
    print("=" * 80)

    for task_id in sorted(tasks.keys()):  # Iterate tasks in ascending ID order
        task_data = tasks[task_id]

        # Skip tasks that lack a 'default' entry: cannot compute pairwise distances
        # without a reference embedding to compare against
        if 'default' not in task_data:
            continue

        # ── Extract default (reference) embedding ─────────────────────────────
        default_emb = task_data['default']['embedding']
        # Flatten to 1D if the extraction script stored a 2D (1, D) array
        if default_emb.ndim > 1:
            default_emb = default_emb.flatten()  # (1, D) → (D,)

        default_cmd      = task_data['default']['command_text']    # Original full instruction
        default_rollouts = task_data['default'].get('num_rollouts', 'N/A')  # Episode count
        default_sr       = task_data['default'].get('success_rate', None)   # Policy success rate

        print(f"\nTask {task_id}")
        # Build success rate suffix: omit entirely if not present in the data
        sr_str = f", SR={default_sr:.0%}" if default_sr is not None else ""
        print(f"  Default: {default_cmd} [{default_rollouts} rollouts{sr_str}]")

        # ── Iterate over each paraphrase level ────────────────────────────────
        for level in ['l1', 'l2', 'l3']:
            if level not in task_data:
                # Level not collected for this task; mark as missing and skip
                print(f"  {level.upper():3s}:     [NOT FOUND]")
                continue

            # ── Extract variant embedding ──────────────────────────────────────
            var_emb = task_data[level]['embedding']
            if var_emb.ndim > 1:
                var_emb = var_emb.flatten()  # Normalise to 1D, same as default_emb

            var_cmd      = task_data[level]['command_text']
            var_rollouts = task_data[level].get('num_rollouts', 'N/A')
            var_sr       = task_data[level].get('success_rate', None)

            # ==================================================================
            # METRIC 1: Cosine Similarity (semantic)
            # cosine_similarity([a], [b]) returns a (1,1) matrix; [0,0] extracts
            # the scalar. Wrapping in lists gives sklearn 2D row-vector inputs.
            # ==================================================================
            cos_sim = cosine_similarity([default_emb], [var_emb])[0, 0]

            # ==================================================================
            # METRIC 2: Euclidean Distance L2 (semantic)
            # euclidean_distances([a], [b]) also returns (1,1); [0,0] for scalar.
            # Measures absolute displacement in the embedding space.
            # ==================================================================
            euc_dist = euclidean_distances([default_emb], [var_emb])[0, 0]

            # ==================================================================
            # METRIC 3: Normalised Levenshtein Distance (lexical)
            # Pure character-level surface-form comparison; independent of
            # the model's representation space.
            # ==================================================================
            lev_dist = compute_levenshtein_normalized(default_cmd, var_cmd)

            # Store all metrics for this (task_id, level) pair as a flat dict.
            # This dict becomes one row in the final DataFrame.
            results.append({
                'task_id':              task_id,
                'level':                level,
                'default_command':      default_cmd,
                'variation_command':    var_cmd,
                'default_rollouts':     default_rollouts,
                'variation_rollouts':   var_rollouts,
                'default_success_rate': default_sr,
                'variation_success_rate': var_sr,
                'cosine_similarity':    cos_sim,
                'euclidean_distance':   euc_dist,
                'levenshtein_distance': lev_dist,
            })

            # ── Per-pair console output ────────────────────────────────────────
            sr_str = f", SR={var_sr:.0%}" if var_sr is not None else ""
            print(f"  {level.upper():3s}:     {var_cmd} [{var_rollouts} rollouts{sr_str}]")
            print(f"           Semantic:   Cosine_sim={cos_sim:.4f}")
            print(f"           Semantic:   Euclidean_dist(L2)={euc_dist:.4f}")
            print(f"           Lexical:    Lev_dist={lev_dist:.4f}")

    # Convert list of dicts to a pandas DataFrame for statistical aggregation
    df = pd.DataFrame(results)

    # ── Per-level aggregate statistics ────────────────────────────────────────
    print("\n" + "=" * 80)
    print("AVERAGE STATISTICS BY LEVEL")
    print("=" * 80)

    for level in ['l1', 'l2', 'l3']:
        # Filter rows belonging to this level
        level_data = df[df['level'] == level]
        if len(level_data) == 0:
            continue  # Skip if no data for this level (e.g. L3 not yet collected)

        print(f"\n{level.upper()}:")
        # Mean ± std for each metric across all tasks at this level
        print(
            f"  Cosine Similarity (semantic):     "
            f"{level_data['cosine_similarity'].mean():.4f} ± "
            f"{level_data['cosine_similarity'].std():.4f}"
        )
        print(
            f"  Euclidean Distance L2 (semantic): "
            f"{level_data['euclidean_distance'].mean():.4f} ± "
            f"{level_data['euclidean_distance'].std():.4f}"
        )
        print(
            f"  Levenshtein Distance (lexical):   "
            f"{level_data['levenshtein_distance'].mean():.4f} ± "
            f"{level_data['levenshtein_distance'].std():.4f}"
        )

    # ── Overall aggregate statistics (all levels combined) ────────────────────
    print("\n" + "=" * 80)
    print("OVERALL STATISTICS")
    print("=" * 80)
    print(
        f"  Overall Cosine Similarity:      "
        f"{df['cosine_similarity'].mean():.4f} ± "
        f"{df['cosine_similarity'].std():.4f}"
    )
    print(
        f"  Overall Euclidean Distance L2:  "
        f"{df['euclidean_distance'].mean():.4f} ± "
        f"{df['euclidean_distance'].std():.4f}"
    )
    print(
        f"  Overall Levenshtein Distance:   "
        f"{df['levenshtein_distance'].mean():.4f} ± "
        f"{df['levenshtein_distance'].std():.4f}"
    )

    # ── Save full results as CSV ───────────────────────────────────────────────
    df.to_csv(output_csv, index=False)  # index=False: omit row-number column
    print(f"\n✓ Results saved: {output_csv}")

    # ── Save per-level formatted Excel tables ─────────────────────────────────
    # Derive the base path by stripping the ".csv" extension.
    # e.g. "/outputs/combined_analysis.csv" → "/outputs/combined_analysis"
    base_path = output_csv.rsplit(".", 1)[0]

    for lvl in ['l1', 'l2', 'l3']:
        level_df = df[df['level'] == lvl].reset_index(drop=True)
        # reset_index ensures the per-level DataFrame has a clean 0-based index
        # so that the alternating row fill logic in _save_formatted_overlap_table
        # works correctly (i % 2 alternation).
        if len(level_df) == 0:
            continue  # Do not create an empty Excel file for missing levels

        formatted_path = f"{base_path}_overlap_{lvl}.xlsx"
        _save_formatted_overlap_table(level_df, lvl, formatted_path)

    return df  # Return DataFrame for programmatic downstream use


# =============================================================================
# EXCEL TABLE FORMATTER (PRIVATE HELPER)
# =============================================================================

def _save_formatted_overlap_table(
    level_df: pd.DataFrame,
    level: str,
    output_path: str,
) -> None:
    """
    Write a human-readable, colour-formatted overlap analysis table to an
    Excel file for a single paraphrase variation level.

    The generated workbook contains one sheet titled ``"{LEVEL} Variations"``
    with the following layout:

    Row 1 — **Title** (merged across all 6 columns, bold, centred)
    Row 2 — **Column headers** (dark blue background, white bold text)
    Rows 3…N+2 — **Data rows** (alternating plain white / light-blue fill)
    Row N+3 — **Average row** (golden fill, bold text, mean ± std format)

    Columns
    -------
    1. N°            — 1-indexed row number
    2. Original Task Command   — default instruction string (wide, wrapped)
    3. Variation Task Command  — variant instruction string (wide, wrapped)
    4. Cosine Similarity       — rounded to 4 decimal places, centred
    5. Euclidean Dist.         — rounded to 4 decimal places, centred
    6. Levenshtein Dist.       — rounded to 4 decimal places, centred

    The Average row shows ``"mean ± std"`` as a formatted string (not a
    formula) so the file is readable without Excel formulas being active.

    Styling Palette
    ---------------
    - Header fill:  ``#2F5496`` (corporate dark blue)
    - Odd data rows: ``#DCE6F1`` (pale blue) — even rows are unfilled (white)
    - Average row:  ``#F4B942`` (amber/gold)
    - All cells:    thin black border on all four sides

    Parameters
    ----------
    level_df : pd.DataFrame
        Filtered DataFrame containing only rows for the target level.
        Must have a clean 0-based integer index (use ``reset_index(drop=True)``
        before calling this function).
    level : str
        Paraphrase level identifier (``"l1"``, ``"l2"``, or ``"l3"``).
        Used in the sheet title and the file-level title row.
    output_path : str
        Absolute path for the ``.xlsx`` output file. Parent directories
        must exist; openpyxl does not create them automatically.

    Returns
    -------
    None
        Saves the workbook to ``output_path`` and prints a confirmation line.
    """
    # Lazy imports: openpyxl is only required for Excel export.
    # Keeping them here avoids an ImportError on systems where openpyxl is not
    # installed but the CSV-only analysis path is used.
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter  # Converts 1→"A", 2→"B", etc.

    # ── Create workbook and active worksheet ──────────────────────────────────
    wb = openpyxl.Workbook()  # New in-memory workbook with one empty sheet
    ws = wb.active            # Reference to the single default sheet
    ws.title = f"{level.upper()} Variations"  # e.g. "L1 Variations"

    # ── Row 1: Title (merged across all 6 columns) ─────────────────────────────
    title = f"Overlapping Analysis - {level.upper()} Variations"
    ws.merge_cells("A1:F1")       # Merge columns A through F in row 1
    title_cell = ws["A1"]
    title_cell.value     = title
    title_cell.font      = Font(bold=True, size=13)               # Bold, 13pt
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    ws.row_dimensions[1].height = 22  # Slightly taller than default for readability

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    headers = [
        "N°",
        "Original Task Command",
        "Variation Task Command",
        "Cosine Similarity",
        "Euclidean Dist.",
        "Levenshtein Dist.",
    ]

    # Dark blue fill for the header row (hex without '#' prefix for openpyxl)
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")

    # White bold font on dark background for contrast
    header_font = Font(bold=True, color="FFFFFF")

    # Thin black border definition, reused for every cell in the table
    thin   = Side(style="thin", color="000000")  # 1px solid black
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):  # Column index starts at 1 (A)
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,  # Allow long header text to wrap within the cell
        )
        cell.border    = border

    ws.row_dimensions[2].height = 20  # Slightly taller header row

    # ── Rows 3 … N+2: Data rows ───────────────────────────────────────────────
    # Alternating fill: odd rows (i % 2 == 1) get a pale-blue background;
    # even rows remain white (no fill applied).
    alt_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")  # Pale blue
    num_fmt_4 = "0.0000"  # 4-decimal-place Excel number format code

    for i, row in level_df.iterrows():
        # Map DataFrame row index to Excel row number (row 0 → Excel row 3)
        excel_row = i + 3

        # Apply alternating fill: odd 0-based index (i=1,3,5,...) = shaded
        fill = alt_fill if i % 2 == 1 else None

        # Build the ordered list of cell values for this row
        values = [
            i + 1,                              # Col 1: 1-indexed row number
            row['default_command'],             # Col 2: original instruction
            row['variation_command'],           # Col 3: variant instruction
            round(row['cosine_similarity'], 4), # Col 4: cosine sim (4 dp)
            round(row['euclidean_distance'], 4),# Col 5: euclidean dist (4 dp)
            round(row['levenshtein_distance'], 4), # Col 6: Levenshtein dist (4 dp)
        ]

        for col, val in enumerate(values, start=1):
            cell           = ws.cell(row=excel_row, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Apply row fill only for odd-indexed rows (shading)
            if fill:
                cell.fill = fill

            # Apply numeric format and centring to the three metric columns (4–6)
            if col >= 4:
                cell.number_format = num_fmt_4
                cell.alignment     = Alignment(
                    horizontal="center",
                    vertical="center",
                )  # Override wrap_text for numeric cells (no wrapping needed)

    # ── Average row ───────────────────────────────────────────────────────────
    avg_row  = len(level_df) + 3   # Immediately below the last data row
    avg_fill = PatternFill(fill_type="solid", fgColor="F4B942")  # Golden amber
    avg_font = Font(bold=True)

    def fmt_mean_std(col_name: str) -> str:
        """Format mean ± std as a string for a numeric DataFrame column."""
        mean = level_df[col_name].mean()
        std  = level_df[col_name].std()
        return f"{mean:.4f} ± {std:.4f}"

    # Build the average row: label in col 1, blanks in cols 2–3, stats in 4–6
    avg_values = [
        "AVERAGE",                             # Col 1: row label
        "",                                    # Col 2: blank (no command average)
        "",                                    # Col 3: blank (no command average)
        fmt_mean_std('cosine_similarity'),     # Col 4: "X.XXXX ± X.XXXX"
        fmt_mean_std('euclidean_distance'),    # Col 5: "X.XXXX ± X.XXXX"
        fmt_mean_std('levenshtein_distance'),  # Col 6: "X.XXXX ± X.XXXX"
    ]

    for col, val in enumerate(avg_values, start=1):
        cell           = ws.cell(row=avg_row, column=col, value=val)
        cell.font      = avg_font
        cell.fill      = avg_fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    # Widths are in Excel "character units" (≈ width of the digit "0").
    col_widths = [6, 52, 52, 18, 16, 18]
    # Col 1 (N°): narrow; Cols 2–3 (commands): wide for long instruction strings;
    # Cols 4–6 (metrics): moderate width for "X.XXXX ± X.XXXX" strings.
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Persist to disk ───────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"✓ Formatted overlap table saved: {output_path}")


# =============================================================================
# PER-ROLLOUT GRANULAR ANALYSIS
# =============================================================================

def compare_per_rollout_similarity(
    embeddings: dict,
    task_id: int = 0,
    level: str = 'l1',
) -> tuple:
    """
    Compute cosine similarity and Euclidean distance between the default and
    variant embeddings **for each individual rollout**, rather than using the
    mean embedding that aggregates all rollouts.

    Motivation
    ----------
    The mean-embedding analysis in ``analyze_embeddings`` provides a single
    representative distance per instruction pair. Per-rollout analysis reveals
    the **variance** in that distance across rollouts — answering questions like:
    "Are there specific episodes where the default and variant embeddings diverge
    sharply, despite being close on average?" High variance suggests that the
    model's internal representation of the two instructions is not consistently
    similar, which may indicate scene-dependent sensitivity to the instruction
    wording.

    Embedding Matching
    ------------------
    Rollout ``i`` of the default instruction is compared against rollout ``i``
    of the variant instruction (1:1 paired matching). If the two have different
    numbers of rollouts, only the first ``min(n_default, n_variant)`` pairs are
    compared. This paired approach controls for scene variation: both embeddings
    come from the same initial simulator state, so differences are attributable
    to the instruction difference rather than the scene.

    Key Lookup Convention
    ---------------------
    Embeddings must be keyed by the pattern ``"task_{task_id:02d}_{level}"``
    (zero-padded to two digits):

    - Default: ``"task_00_default"``, ``"task_07_default"``, etc.
    - Variant: ``"task_00_l1"``, ``"task_07_l2"``, etc.

    Parameters
    ----------
    embeddings : dict[str, dict]
        Pre-loaded embedding dictionary from ``load_embeddings`` or
        ``pickle.load``. Must contain ``"embedding_per_rollout"`` arrays
        for both the default and the target variant level.
    task_id : int, default ``0``
        0-indexed LIBERO-Goal task ID to analyse.
    level : str, default ``"l1"``
        Paraphrase level to compare against default.
        Must be one of ``"l1"``, ``"l2"``, ``"l3"``.

    Returns
    -------
    similarities : np.ndarray of shape ``(min_rollouts,)``
        Cosine similarity for each paired rollout. Values in ``[-1, 1]``.
        Returns ``None`` if keys are missing or per-rollout data is absent.
    euc_dists : np.ndarray of shape ``(min_rollouts,)``
        Euclidean (L2) distance for each paired rollout. Non-negative.
        Returns ``None`` if keys are missing or per-rollout data is absent.

    Notes
    -----
    When both return values are ``None`` (error conditions), the function also
    prints a human-readable reason to stdout and returns ``None`` (not a tuple).
    Callers should guard against ``None`` returns.
    """
    # ── Build lookup keys from task_id and level ───────────────────────────────
    default_key = f"task_{task_id:02d}_default"  # e.g. "task_07_default"
    var_key     = f"task_{task_id:02d}_{level}"  # e.g. "task_07_l1"

    # ── Guard: check both keys exist in the loaded embeddings ─────────────────
    if default_key not in embeddings or var_key not in embeddings:
        print(f"Keys not found: {default_key} or {var_key}")
        return None  # Cannot compare; caller should handle None gracefully

    default_data = embeddings[default_key]  # Full metadata dict for the default
    var_data     = embeddings[var_key]      # Full metadata dict for the variant

    # ── Guard: check per-rollout arrays exist in both entries ─────────────────
    if ('embedding_per_rollout' not in default_data
            or 'embedding_per_rollout' not in var_data):
        print("Per-rollout embeddings not available")
        return None  # Data was collected in mean-only mode; skip

    # ── Extract per-rollout embedding matrices ─────────────────────────────────
    default_rollouts = default_data['embedding_per_rollout']  # (R_d, D)
    var_rollouts     = var_data['embedding_per_rollout']      # (R_v, D)

    # Use the shorter of the two to avoid index-out-of-bounds errors when
    # the two collections have different rollout counts
    min_rollouts = min(len(default_rollouts), len(var_rollouts))

    # ── Per-rollout cosine similarity ─────────────────────────────────────────
    similarities = []
    for i in range(min_rollouts):
        # Compare rollout i of the default against rollout i of the variant.
        # cosine_similarity returns a (1,1) matrix; [0,0] extracts the scalar.
        sim = cosine_similarity(
            [default_rollouts[i]],  # Single row vector wrapped in a list
            [var_rollouts[i]],      # Single row vector wrapped in a list
        )[0, 0]
        similarities.append(sim)

    similarities = np.array(similarities)  # Convert list to NumPy for stats

    # ── Per-rollout Euclidean distance ─────────────────────────────────────────
    euc_dists = []
    for i in range(min_rollouts):
        # euclidean_distances returns a (1,1) matrix; [0,0] extracts the scalar.
        dist = euclidean_distances(
            [default_rollouts[i]],
            [var_rollouts[i]],
        )[0, 0]
        euc_dists.append(dist)

    euc_dists = np.array(euc_dists)  # Convert list to NumPy for stats

    # ── Console report ────────────────────────────────────────────────────────
    print(f"\nTask {task_id} - Default vs {level.upper()}")
    print(f"  Default command:   {default_data['command_text']}")
    print(f"  Variation command: {var_data['command_text']}")

    print(f"\n  Per-rollout cosine similarity ({min_rollouts} rollouts):")
    print(f"    Mean:   {similarities.mean():.4f}")  # Central tendency
    print(f"    Std:    {similarities.std():.4f}")   # Spread across rollouts
    print(f"    Min:    {similarities.min():.4f}")   # Worst-case rollout
    print(f"    Max:    {similarities.max():.4f}")   # Best-case rollout

    print(f"\n  Per-rollout euclidean distance L2 ({min_rollouts} rollouts):")
    print(f"    Mean:   {euc_dists.mean():.4f}")
    print(f"    Std:    {euc_dists.std():.4f}")
    print(f"    Min:    {euc_dists.min():.4f}")
    print(f"    Max:    {euc_dists.max():.4f}")

    return similarities, euc_dists


# =============================================================================
# COMMAND-LINE INTERFACE
# =============================================================================

if __name__ == "__main__":
    import argparse  # Standard-library CLI argument parser

    # ── Argument parser setup ─────────────────────────────────────────────────
    parser = argparse.ArgumentParser(
        description="Analyze command variations with rollout-based embeddings",
        formatter_class=argparse.RawDescriptionHelpFormatter,  # Preserve whitespace
                                                                # in the epilog examples
        epilog="""
Examples:
  # Single file
  python analyze_embedding_overlap_rollout.py --embedding_file embeddings.pkl

  # Multiple files (one per command level)
  python analyze_embedding_overlap_rollout.py --embedding_files \\
      /home/A.CARDAMONE7/outputs/embeddings/rollout_embeddings_libero_goal_default_first_step_r10.pkl \\
      /home/A.CARDAMONE7/outputs/embeddings/rollout_embeddings_libero_goal_l1_first_step_r10.pkl \\
      /home/A.CARDAMONE7/outputs/embeddings/rollout_embeddings_libero_goal_l2_first_step_r10.pkl \\
      /home/A.CARDAMONE7/outputs/embeddings/rollout_embeddings_libero_goal_l3_first_step_r10.pkl
        """,
    )

    # ── Embedding source arguments (mutually exclusive in practice) ────────────
    parser.add_argument(
        "--embedding_file",
        type=str,
        default=None,
        help="Path to a single embeddings pickle file",
        # Single-file mode: backwards-compatible with older scripts that
        # stored all levels in one .pkl
    )
    parser.add_argument(
        "--embedding_files",
        type=str,
        nargs="+",   # Accept one or more space-separated paths
        default=None,
        help="Paths to multiple embeddings pickle files (one per command level)",
    )
    parser.add_argument(
        "--embedding_dir",
        type=str,
        default=None,
        help="Directory containing embeddings pickle files",
        # All *.pkl files found in this directory are loaded and merged
    )

    # ── Output path ────────────────────────────────────────────────────────────
    parser.add_argument(
        "--output_csv",
        type=str,
        default=None,
        help="Output CSV path (default: auto-generated from input path)",
    )

    # ── Per-rollout analysis flags ─────────────────────────────────────────────
    parser.add_argument(
        "--per_rollout_analysis",
        action="store_true",  # Flag: True if present, False if absent
        help="Run per-rollout similarity analysis for a specific task",
    )
    parser.add_argument(
        "--task_id",
        type=int,
        default=0,
        help="Task ID for per-rollout analysis (0-indexed)",
    )
    parser.add_argument(
        "--level",
        type=str,
        default="l1",
        help="Command level for per-rollout analysis (l1, l2, or l3)",
    )

    args = parser.parse_args()  # Parse sys.argv into the args namespace

    # ── Embedding loading: three mutually exclusive modes ─────────────────────
    if args.embedding_file:
        # ── Mode 1: Single file (backwards-compatible) ────────────────────────
        # Directly open and unpickle the file without going through load_embeddings.
        with open(args.embedding_file, 'rb') as f:
            embeddings = pickle.load(f)
        print(f"Loaded {len(embeddings)} embeddings from {args.embedding_file}")

        # Output CSV sits alongside the input file (same directory, same stem)
        output_base = args.embedding_file

    elif args.embedding_files or args.embedding_dir:
        # ── Mode 2: Multiple files or directory ───────────────────────────────
        # Delegate to load_embeddings which handles sorting and merging.
        embeddings = load_embeddings(
            embedding_files=args.embedding_files,
            embedding_dir=args.embedding_dir,
        )

        # Derive the output base path from the directory or the first file's parent
        output_base = (
            args.embedding_dir
            or os.path.dirname(args.embedding_files[0])  # Parent dir of first file
        )
        output_base = os.path.join(output_base, "combined_analysis")
        # e.g. "/outputs/embeddings/combined_analysis" → CSV/XLSX go here

    else:
        # ── Mode 3: Default fallback (no CLI args provided) ───────────────────
        # Try loading from the hard-coded cluster storage path. Useful when
        # running interactively without CLI arguments during development.
        default_dir = "/mnt/beegfs/a.cardamone7/outputs/embeddings/openvla/"
        embeddings  = load_embeddings(embedding_dir=default_dir)
        output_base = os.path.join(default_dir, "combined_analysis")

    # ── Run main analysis ─────────────────────────────────────────────────────
    # Use the user-specified output CSV path if provided; otherwise auto-derive
    # it by appending ".csv" to the computed output_base path.
    df = analyze_embeddings(
        embeddings,
        output_csv=args.output_csv or f"{output_base}.csv",
    )

    # ── Optional per-rollout granular analysis ────────────────────────────────
    # Only executed when --per_rollout_analysis flag is present.
    if args.per_rollout_analysis:
        similarities, euc_dists = compare_per_rollout_similarity(
            embeddings,
            task_id=args.task_id,  # 0-indexed task ID from --task_id
            level=args.level,      # Level string from --level
        )
        # Results are printed inside compare_per_rollout_similarity;
        # returned arrays can be used interactively in a REPL or notebook.